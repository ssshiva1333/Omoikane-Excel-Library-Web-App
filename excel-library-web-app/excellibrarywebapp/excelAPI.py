import shutil
import datetime
from pathlib import Path
from openpyxl import load_workbook

class excelAPI:
    current_directory = Path(__file__).resolve().parent
    excel_directory = current_directory.parent / 'excel'
    excel_file = excel_directory / 'books.xlsx'

    @staticmethod
    def addBook(data):
        try:
            wb = load_workbook(excelAPI.excel_file)
            ws = wb.active

            headers = [cell.value for cell in ws[1]]
            if not headers:
                raise ValueError("No headers found in the Excel sheet.")

            next_row = ws.max_row + 1

            for key in ['bName', 'aName', 'pNumber', 'genre', 'eNumber', 'pName']:
                if key in headers:
                    col_index = headers.index(key) + 1
                    value = data.get(key)
                    if value:
                        print(f"Writing '{value}' to column {col_index}")
                        ws.cell(row=next_row, column=col_index, value=value)
                    else:
                        print(f"No value for '{key}'")
                else:
                    print(f"Warning: '{key}' not found in headers.")

            wb.save(excelAPI.excel_file)
        except Exception as e:
            print(f"Error in addBook: {e}")

    @staticmethod
    def getBook(data={}):
        try:
            wb = load_workbook(excelAPI.excel_file)
            ws = wb.active

            headers = [cell.value for cell in ws[1]]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            data_list = [dict(zip(headers, row)) for row in rows]

            if not data:
                return data_list

            filtered_data = []
            for row in data_list:
                match = True
                for key, value in data.items():
                    if value is None:
                        continue
                    
                    if key in row and row[key] is not None:
                        if isinstance(value, str):
                            if value.lower() not in str(row[key]).lower():
                                match = False
                                break
                        elif row[key] != value:
                            match = False
                            break

                if match:
                    filtered_data.append(row)

            return filtered_data

        except Exception as e:
            print(f"Error in getBook: {e}")
            return [{}]
        
    @staticmethod
    def changeBook(data):
        try:
            # Load workbook and select the active sheet
            wb = load_workbook(excelAPI.excel_file)
            ws = wb.active

            # Read headers from the first row
            headers = [cell.value for cell in ws[1]]
            if 'bName' not in headers:
                raise ValueError("'bName' column not found in headers.")

            # Iterate through the rows to find the one with matching book name
            book_name = data.get('bName')
            if not book_name:
                raise ValueError("No 'bName' provided in input data.")
            
            row_found = False
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False), start=2):
                # Compare the value in the 'bName' column with the provided book name
                if row[headers.index('bName')].value == book_name:
                    row_found = True
                    # Update the row with the provided data
                    for key, value in data.items():
                        if key in headers and value is not None:
                            col_index = headers.index(key) + 1
                            ws.cell(row=row_idx, column=col_index, value=value)
                    break

            if not row_found:
                print(f"Book with name '{book_name}' not found.")
            else:
                # Save workbook after updating
                wb.save(excelAPI.excel_file)
                print(f"Book '{book_name}' updated successfully.")

        except Exception as e:
            print(f"Error in updateBook: {e}")

    @staticmethod
    def removeBook(data):
        try:
            wb = load_workbook(excelAPI.excel_file)
            ws = wb.active

            headers = [cell.value for cell in ws[1]]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            
            # Convert rows to dictionaries
            data_list = [dict(zip(headers, row)) for row in rows]

            # Filter rows based on criteria
            filtered_rows = []
            for row in data_list:
                match = True
                for key, value in data.items():
                    if key in row:
                        if value:  # If a value is provided for the key
                            if row[key] != value:
                                match = False
                                break
                    else:  # If no value is provided, skip the criteria check
                        continue

                # Add rows that do not match the criteria to filtered_rows
                if not match:
                    filtered_rows.append(row)

            # Clear existing rows and write filtered rows
            ws.delete_rows(2, ws.max_row - 1)  # Adjust the range to avoid deleting the header
            for row_index, row in enumerate(filtered_rows, start=2):
                for col_index, value in enumerate(row.values(), start=1):
                    ws.cell(row=row_index, column=col_index, value=value)

            wb.save(excelAPI.excel_file)

        except Exception as e:
            print(f"Error in removeBook: {e}")

    @staticmethod
    def excelBackup():
        try:
            backup_filename = f"books-backup-{datetime.date.today()}.xlsx"
            backup_file = excelAPI.excel_directory / backup_filename
            
            shutil.copy(excelAPI.excel_file, backup_file)

        except Exception as e:
            print(f"Error in excel backup: {e}")